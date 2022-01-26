# *********************************************
# 아파트 시세조회 프로그
# 작성자 : 장홍진
# 작성일 : 2021.08.08 (ver 1.0)
# 주요기능 
#   1.한국 부동산원의 아파트 시세 조회
#   2.KB 국민은행 아파트 시세 조회
# 문의사항 : hazelnuttree@naver.com

# 업데이트 이력
# v1.1 (2021.08.12)
#  -  kb 부동산 정보 조회를 위한 행정동 코드 업데이트
#  -  부동산 목록이 없는 경우 콤보창에 관련 메세지 노출

# v2.0 (2021.08.15)
#  -  엑셀 출력 자동 설정
#  -  국토부 실거래가 조회 추가

import tkinter as tk
import tkinter.ttk
import requests
import openpyxl
import datetime
import time
import sys
import pandas as pd
import numpy as np
import xlsxwriter
import easygui
import random
import win32com.client as win32

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from tkinter import *
from tkinter import messagebox
from PyQt5.QtWidgets import QComboBox, QWidget, QLabel, QApplication, QPushButton, QCheckBox
from PyQt5.QtCore import QCoreApplication

from bs4 import BeautifulSoup
from urllib import request as req
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

import json
import sys

import os
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = '/path/to/Anaconda3/Library/plugins/platforms'

class UI_FORM(QWidget):
    
    def __init__(self):

        super().__init__()
        
        self.initUI()

        sel_do = Select(driver.find_element_by_id('do_code1'))    
        
        list_do = self.list_of_do(sel_do)   

        self.Combo1(list_do)
            
    def initUI(self):    
        
        # 콤보박스
        self.cb1 = QComboBox(self)
        self.cb1.setGeometry(50,80,100,25)
        
        self.cb2 = QComboBox(self)
        self.cb2.setGeometry(200,80,100,25)

        self.cb3 = QComboBox(self)
        self.cb3.setGeometry(350,80,100,25)
     
        self.cb4 = QComboBox(self)
        self.cb4.setGeometry(500,80,350,25)

        self.cb5 = QComboBox(self)
        self.cb5.setGeometry(500,210,350,25)
        self.cb5.setEnabled(False)        

        self.cb6 = QComboBox(self)
        self.cb6.setGeometry(500, 380, 350, 25)
        self.cb6.setEnabled(False)

        self.cb7 = QComboBox(self)
        self.cb7.setGeometry(530, 350, 50, 20)
        self.cb7.setEnabled(False)
        list_year = ['2019', '2020', '2021', '2022']
        self.cb7.addItems(list_year)
        now_year = datetime.datetime.now().year
        str_now_year = str(now_year)
        if str_now_year not in list_year:
            self.cb7.addItem(str_now_year)
        self.cb7.setCurrentText(str_now_year)
        self.cb7.activated.connect(self.comboBoxFunction7)

        self.cb7_lb = QLabel('년도', self)
        self.cb7_lb.setFixedWidth(30)
        self.cb7_lb.move(500, 355)

        self.cb8 = QComboBox(self)
        self.cb8.setGeometry(630, 350, 30, 20)
        self.cb8.setEnabled(False)
        list_qrt = ['1', '2', '3', '4']
        self.cb8.addItems(list_qrt)
        now_quarter = (datetime.datetime.now().month - 1) // 3 + 1
        self.cb8.setCurrentText(str(now_quarter))
        self.cb8.activated.connect(self.comboBoxFunction8)

        self.cb8_lb = QLabel('분기', self)
        self.cb8_lb.setFixedWidth(30)
        self.cb8_lb.move(600, 355)

        self.cb8_chk = QCheckBox('', self)
        self.cb8_chk.setFixedWidth(30)
        self.cb8_chk.move(670, 355)
        self.cb8_chk.stateChanged.connect(self.checkBoxFunction)

        # 버튼
        self.btn1 = QPushButton('한국부동산원',self)
        self.btn1.move(630,120)
        self.btn1.setEnabled(False)        
        self.btn1_state = 0
        self.btn1.clicked.connect(self.btn1_clicked)

        self.btn1_guide = QLabel('', self)        
        self.btn1_guide.setFixedWidth(50)        
        self.btn1_guide.move(580,123)

        self.btn2 = QPushButton('KB 부동산',self)
        self.btn2.move(630,250)
        self.btn2.setEnabled(False)
        self.btn1_state = 0
        self.btn2.clicked.connect(self.btn2_clicked)

        self.btn2_guide = QLabel('', self)        
        self.btn2_guide.setFixedWidth(50)        
        self.btn2_guide.move(580,253)
        
        self.btn3 = QPushButton('실거래가',self)
        self.btn3.move(630,420)
        self.btn3.setEnabled(False)
        self.btn3.clicked.connect(self.btn3_clicked)

        self.btn3_guide = QLabel('', self)        
        self.btn3_guide.setFixedWidth(50)        
        self.btn3_guide.move(580,423)

        self.btn4 = QPushButton('재시작',self)
        self.btn4.move(50,380)
        self.btn4.setEnabled(True)
        self.btn4.clicked.connect(self.btn4_clicked)

        self.btn5 = QPushButton('종료',self)
        self.btn5.move(150,380)
        self.btn5.setEnabled(True)
        self.btn5.clicked.connect(self.btn5_clicked)

        # 라벨박스
        self.lb1 = QLabel('가.시 도 선택', self)
        self.lb1.setFixedWidth(110)        
        self.lb1.move(50,60)
        
        self.lb2 = QLabel('나.구 군 선택', self)
        self.lb2.setFixedWidth(110)        
        self.lb2.move(200,60)

        self.lb3 = QLabel('다.읍면동 선택', self)
        self.lb3.setFixedWidth(110)        
        self.lb3.move(350,60)

        self.lb4 = QLabel('라. 아파트 선택 : 한국부동산원', self)
        self.lb4.setFixedWidth(450)        
        self.lb4.move(500,60)

        self.lb5 = QLabel('마. 아파트 선택 : KB 국민은행', self)
        self.lb5.setFixedWidth(450)
        self.lb5.move(500,190)

        self.lb6 = QLabel('바. 아파트 선택 : 국토교통부 실거래가', self)
        self.lb6.setFixedWidth(450)
        self.lb6.move(500,320)

        self.msg1 = QLabel('＃ 한국부동산원 처리 후 아래 KB 부동산 선택이 가능합니다', self)
        self.msg1.setFixedWidth(450)
        self.msg1.move(500,160)

        self.msg2 = QLabel('＃ KB 부동산 처리 후 아래 국토부 실거래가 선택이 가능합니다', self)
        self.msg2.setFixedWidth(450)
        self.msg2.move(500,290)

        self.msg3 = QLabel('＃ 아직 처리가 완료되지 않았습니다', self)
        self.msg3.setFixedWidth(450)
        self.msg3.move(500,460)

        self.head1 = QLabel('신용보증기금 구상실익 검토를 위한 아파트 시세조회 프로그램   (※ 처리 속도가 빠르지 않으니 "천천히" 진행해 주시기 바랍니다. ^^;;)', self)
        self.head1.move(50, 20)

        self.head2 = QLabel('========================================================', self)
        self.head2.move(50, 30)

        self.body1 = QLabel('♧  I LOVE KODIT  ♧', self)
        self.body1.move(50,160)
        font1=self.body1.font()
        font1.setPointSize(10)
        font1.setFamily('Times New Roman')
        font1.setBold(True)
        self.body1.setFont(font1)

        self.body2 = QLabel('', self)
        self.body2.setFixedWidth(450)
        self.body2.move(200,160)

        self.body3 = QLabel('※ Process : ', self)
        self.body3.move(50,210)

        self.body4 = QLabel('Welcome', self)
        self.body4.setFixedWidth(450)
        self.body4.move(130, 210)

        self.body5 = QLabel(" ('■' 6개가 모일 때까지 진행해 주세요)", self)
        self.body5.setFixedWidth(450)
        self.body5.move(115, 230)

        self.tail = QLabel('※ 문의처 : hazelnut@kodit.co.kr    (유효기간 2022.12.31)', self)
        self.tail.setFixedWidth(450)
        self.tail.move(50,460)

        self.setWindowTitle('KODIT APT PRICE v2.0')
        self.setGeometry(200,200,900,500)

        self.show()

    # 가. 한국부동산원 아파트 시세

    def Combo1(self, list_do):
        self.cb1.clear()
        self.cb2.clear()
        self.cb3.clear()
        self.cb4.clear()

        msg_list = ['환영합니다', '반갑습니다', '업무에 도움이 되었으면 좋겠어요', '신보의 주인공은 당신이예요', '어제보다 나은 오늘이죠?', '이 또한 지나갈거예요 파이팅!!',
                    '열심히 일하는 당신! 아름다워요']
        msg = random.choice(msg_list)
        self.body2.setText(str(msg))

        self.body4.setText('□□□□□□')

        self.cb1.addItems(list_do)

        self.cb1.activated.connect(self.comboBoxFunction1)

    def Combo2(self, list_city):
        self.cb2.clear()
        self.cb3.clear()
        self.cb4.clear()

        self.cb2.addItems(list_city)
        self.cb2.activated.connect(self.comboBoxFunction2)

    def Combo3(self, list_dong):
        self.cb3.clear()
        self.cb4.clear()

        self.cb3.addItems(list_dong)
        self.cb3.activated.connect(self.comboBoxFunction3)

    def Combo4(self, list_apt):
        self.cb4.clear()

        self.cb4.addItems(list_apt)
        self.cb4.activated.connect(self.comboBoxFunction4)

    def Combo5(self, list_kb):
        self.cb5.clear()

        list_kb_key = list(list_kb.keys())
        self.cb5.addItems(list_kb_key)
        self.cb5.activated.connect(self.comboBoxFunction5)

    def Combo6(self, list_rt):
        self.cb6.clear()

        self.cb6.addItems(list_rt)
        self.cb6.activated.connect(self.comboBoxFunction6)

    # def Combo7(self):
    #     self.cb7.clear()
    #
    #     list_year = ['2019','2020','2021','2022']
    #     self.cb7.addItems(list_year)
    #     print('x1')
    #     now_year = datetime.datetime.now().yaer
    #     self.cb7.setCurrentText(str(now_year))
    #     self.cb7.activated.connect(self.comboBoxFunction7)
    #
    # def Combo8(self):
    #     self.cb8.clear()
    #
    #     list_qrt = ['1', '2', '3', '4']
    #     self.cb8.addItems(list_qrt)
    #     print('x2')
    #     now_quarter = (datetime.datetime.now().month - 1) // 3 + 1
    #     print('now_quarter')
    #     self.cb8.setCurrentText(str(now_quarter))
    #     self.cb8.activated.connect(self.comboBoxFunction8)

    # 시도 이름 구하기
    def list_of_do(self, sel_do):
        global list_do
        list_do=[]
        i1 = 0

        list_do.append('선택')
        for i1 in range(2, len(sel_do.options)+1):
            do = driver.find_element_by_xpath('//*[@id="do_code1"]/option[' + str(i1) + ']')    
            do_name = do.text
            # 시도 리스트 작성
            list_do.append(do_name)
        print('시도 리스트 OK')        
        return list_do
        
    # 구군 이름 구하기
    def list_of_city(self, sel_city):
        global list_city
        list_city=[]
        i2 = 0

        list_city.append('선택')
        for i2 in range(2, len(sel_city.options)+1):
            city = driver.find_element_by_xpath('//*[@id="city_code1"]/option[' + str(i2) + ']')    
            city_name = city.text
            # 시도 리스트 작성
            list_city.append(city_name)
        print('구군 리스트 OK')                    
        return list_city

    # 읍면동 이름 구하기
    def list_of_dong(self, sel_dong):
        global list_dong
        list_dong=[]
        i3=0

        list_dong.append('선택')
        for i3 in range(2, len(sel_dong.options)+1):
            dong = driver.find_element_by_xpath('//*[@id="dong_code1"]/option[' + str(i3) + ']')    
            dong_name = dong.text
            # 동 리스트 작성
            list_dong.append(dong_name)
        print('읍면동 리스트 OK')                    
        return list_dong

    # 아파트 이름 구하기
    def list_of_apt(self, sel_apt):
        global list_apt
        list_apt=[]
        i4 = 0

        list_apt.append('선택')
        for i4, val4 in enumerate(sel_apt):
            
            # 아파트 이름 추출
            apt_name=str(val4.text.strip())

            # 아파트 정보가 없는 경우
            if apt_name == "단지 정보가 없습니다.":
                list_apt[0] = "아파트 정보 없음"
                print("break")
                break

            # 아파트 정보가 있는 경우    
            else :            
                if apt_name[0] == "아":
                    apt_name_pre = apt_name[0] + '파트 : '
                    apt_name2 = apt_name_pre + apt_name[1:]
                
                elif apt_name[0] == "오":
                    apt_name_pre = apt_name[0] + '피스텔 : '
                    apt_name2 = apt_name_pre + apt_name[1:]            
                    
                else:
                    apt_name_pre = '기타 : '
                    apt_name2 = apt_name_pre + apt_name[1:]

                # 아파트 리스트 작성    
                list_apt.append(apt_name2)     
        
        print('아파트 리스트 OK')   

        return list_apt

    # 나. KB 부동산 시세

    # KB 부동산 아파트 이름 구하기

    # 에러처리용A            
    def kb(self, dong_code):
        # print('dong_code',dong_code)        

    # 에러처리용B        
    # def kb(self):

        print('KB 부동산 진입')
        
        quest = '법정동코드='
        
        # 에러처리용A
        answer = dong_code

        # 에러처리용B
        # answer = '1168010300'

        # 해당 지역 아파트 정보 추출
        url_kb1 = 'https://api.kbland.kr/land-complex/complexComm/hscmList?' + str(quest) + str(answer)
        
        dict_data = {'key' : 'value'}
        response_kb1 = requests.get(url=url_kb1, params=dict_data)
        json_data1 = json.loads(response_kb1.text)

        self.list_kb = {}

        # 아파트 정보가 없는 경우
        if json_data1['dataBody']['resultCode'] == 33210 :
            self.list_kb['아파트 정보 없음'] = ['', '', '']
        
        # 아파트 정보가 있는 경우            
        else:
            self.list_kb['아파트 선택'] = ['단지일련번호', '위도', '경도']
            i6 = 0
            for i6 in range(len(json_data1['dataBody']['data'])):
    
                list_kb_name = json_data1['dataBody']['data'][i6]['단지명']
                list_sno_name = json_data1['dataBody']['data'][i6]['단지기본일련번호']
                list_x_name = json_data1['dataBody']['data'][i6]['wgs84위도']
                list_y_name = json_data1['dataBody']['data'][i6]['wgs84경도']
                
                self.list_kb[list_kb_name] = [list_sno_name,list_x_name,list_y_name]

            print('KB 부동산 리스트 OK')

        # 아파트 리스트 전달
        self.Combo5(self.list_kb)

    # 다. 부동산 실거래가 조회
    def rt(self, si_code, gu_code, dong_code):
        print('RT 부동산 진입')
        # print(si_code, gu_code, dong_code)
        # print('dong_code',dong_code)

        rt_url1 = 'https://rt.molit.go.kr/srh/getDanjiComboAjax.do'

        print('year', self.year)

        # 왜 안되는지 연구 필요 (더 좋은 로직)
        # if self.year is None:
        #     print('x2')
        #     self.year = datetime.datetime.now().year
        #
        # if self.quarter is None:
        #     self.quarter = (datetime.datetime.now().month - 1) // 3 + 1

        rt_post_data1 = {
            'menuGubun': 'A',
            'srhType': 'LOC',
            'houseType': 1,
            'srhYear': self.year,
            'srhPeriod': self.quarter,
            'gubunCode': 'LAND',
            'sidoCode': si_code,
            'gugunCode': gu_code,
            'dongCode': dong_code,
            'rentAmtType': 3
        }

        rt_res1 = requests.post(rt_url1, rt_post_data1)
        rt_json_data1 = json.loads(rt_res1.text)

        self.list_rt = {}
        self.list_rt['아파트 선택'] =['단지 일련번호']
        i8 = 0
        for i8 in range(len(rt_json_data1['jsonList'])):
            rt_apt_name = rt_json_data1['jsonList'][i8]['NAME']
            rt_apt_cd = rt_json_data1['jsonList'][i8]['CODE']
            self.list_rt[rt_apt_name] = [rt_apt_cd]
        print('RT 부동산 리스트 OK')

        # 아파트 리스트 전달
        self.Combo6(self.list_rt)

    def comboBoxFunction1(self):
        global text1
        list_city=''
        idx1 = self.cb1.currentIndex()
        text1 = self.cb1.currentText()
        print('시도 선택 OK : ', text1)                                                
        self.cb1.setEnabled(False)                                                                
        self.body4.setText("■□□□□□")
        
        if text1 == "선택":
            self.cb2.clear()
            self.cb3.clear()
            self.cb4.clear()
        else : 
            self.lb1.setText( "■ : " + str(text1) + " 선택완료")
            self.lb1.adjustSize()        
    
            path_way = '//*[@id="do_code1"]/option[' + str(idx1+1) + ']'
            driver.find_element_by_xpath(path_way).click()  
    
            sel_city = Select(driver.find_element_by_id('city_code1'))     
            list_city = self.list_of_city(sel_city)     

            # 도시명 전달            
            self.Combo2(list_city)
        
    def comboBoxFunction2(self):
        global text2
        list_dong=''
        idx2 = self.cb2.currentIndex()
        text2 = self.cb2.currentText()
        print('구군 선택 OK : ', text2)
        self.cb2.setEnabled(False)                                                        
        self.body4.setText("■■□□□□")
        
        if text2 == "선택":
            self.cb3.clear()
            self.cb4.clear()
        else:
            self.lb2.setText( "■ : " + str(text2) + " 선택완료")
            self.lb2.adjustSize()        
    
            path_way = '//*[@id="city_code1"]/option[' + str(idx2+1) + ']'
            driver.find_element_by_xpath(path_way).click()  
    
            sel_dong = Select(driver.find_element_by_id('dong_code1'))
            list_dong = self.list_of_dong(sel_dong)        
            
            # 동명 전달
            self.Combo3(list_dong)

    def comboBoxFunction3(self):
        global text3, dong_code
        list_apt=''
        idx3 = self.cb3.currentIndex()
        text3 = self.cb3.currentText()
        print('읍면동 선택 OK : ', text3)          
        self.cb3.setEnabled(False)                                                                                                      
        self.body4.setText("■■■□□□")

        if text3 == "선택":
            self.cb4.clear()
        else:            
            self.lb3.setText( "■ : " + str(text3) + " 선택완료")
            self.lb3.adjustSize()

            try:
                path_way = '//*[@id="dong_code1"]/option[' + str(idx3+1) + ']'

                # KB 전달을 위한 동코드 생성
                dong_code_css = driver.find_element_by_xpath(path_way)
                dong_code = dong_code_css.get_attribute("value")

                # 아파트 리스트 탐색
                driver.find_element_by_xpath(path_way).click()
                driver.find_element_by_class_name('map_search_inputtxt2_search2').click()

                sel_apt = driver.find_elements_by_css_selector('#aptListArea a')

                # 에러처리용A : 아파트 리스트 생성
                list_apt = self.list_of_apt(sel_apt)

                # 에러처리용B
                # list_apt = ['선택','개포6차우성', '개포우성3차', '개포자이']
                
            except:
                easygui.msgbox("시스템 에러가 발생했습니다.(cbf3) 다시 시작해 주세요", title="Error Message")

            self.lb4.setText("(주의!!) 선택 후 '한국부동산원'이 활성화 될때까지 기다려 주세요")

            # 아파트 리스트 전달
            self.Combo4(list_apt)
            
            # KB 콤보박스 활성화
            self.cb5.setEnabled(True)        

    def comboBoxFunction4(self):
        global si_code, gu_code
        idx4 = self.cb4.currentIndex()
        text4 = self.cb4.currentText()
        print('아파트 선택 OK : ', text4)                                                        
        
        # 아파트 특성 정보 구하기        
        if idx4 > 0:

            try : 
                i = idx4 - 1
                
                # 아파트 세부 정보 수집을 위한 팝업 하이퍼링크 생성
                aptlist = driver.find_element_by_xpath('//*[@id="aptListArea"]')
                apt_link_href = [elem.get_attribute("href") for elem in aptlist.find_elements_by_tag_name('a')]
                apt_link_text = str(apt_link_href[i]).strip()

                apt_link_text = apt_link_text.replace("javascript:go_apt_info(",'')
                apt_link_text = apt_link_text.replace(");",'')       
                apt_link_text = apt_link_text.replace("\'",'')                                      
                apt_link_split = apt_link_text.split(',')
                
                link_1 = int(apt_link_split[0])
                link_2 = int(apt_link_split[1])
                link_3 = int(apt_link_split[2])

                # 하이퍼링크 변수 추출                        
                msg_box = str(link_1) + '-' + str(link_2) + '-' + str(link_3)
                # print(msg_box)
                si_code = link_1
                gu_code = link_2

            except :
                easygui.msgbox("시스템 에러가 발생했습니다.(cbf4) 다시 시작해 주세요", title="Error Message")
        
            # 아파트 세부 정보 팝업 하이퍼링크 경로
            popup = 'http://www.rtech.or.kr/rtech/MarketPrice/getMarketPriceDetail.do?categoryCd=' + str(link_1) + '&aptSeq=' + str(link_2) + '&addrCode=' + str(link_3) + '&price_info=1'
            
            driver2.get(popup)
            
            print('아파트 세부 정보 POPUP OK')                                                        
            
            # 브라우저 창크기 조절
            # driver1.set_window_size(200,200)            
            # driver.close()
            
            self.btn1.setEnabled(True)
            self.btn1_guide.setText("클릭 ☞")

        else :
            pass

    def comboBoxFunction5(self):

        idx5 = self.cb5.currentIndex()
        kb_apt = self.cb5.currentText()
        print('KB 부동산 선택 OK : ', kb_apt)

        if idx5 > 0:
            try:
                self.lb5.setText('☞ KB 부동산 제공 아파트"를 선택해 주세요 ')

                # 스크린샷을 위한 주소
                info = self.list_kb[kb_apt]

                sno = info[0] #아파트 고유번호
                x = info[1]
                y = info[2]

                url2 = 'https://kbland.kr/c/' + str(sno) + '?ctype= 01&xy=' + str(x) + ',' + str(y) + ',17'
                driver3.get(url2)

                print('KB 부동산 지도 정보 OK')

                # 아파트 세부정보 이동
                url_kb2 = 'https://api.kbland.kr/land-complex/complex/typInfo?단지기본일련번호=' + str(sno)

                print('KB 부동산 세부 정보 POPUP OK')

                dict_data = {'key' : 'value'}
                response_kb2 = requests.get(url=url_kb2, params=dict_data)
                json_data2 = json.loads(response_kb2.text)

                # 아파트 면적 일련번호 구하기
                list_area = []
                i7=0
                for i7 in range(len(json_data2['dataBody']['data'])):
                    area_sno_name = json_data2['dataBody']['data'][i7]['면적일련번호']
                    list_area.append(area_sno_name)

                area_sno = None
                price_info =[]
                price_info.append(['공급면적평수','공급면적','전용면적','매매일반거래가','매매상한가','매매하한가','매매평균가','기준년월일'])

                # 아파트 시세정보 구하기
                for idx, area_sno in enumerate(list_area):

                    # 아파트 평형별 정보 수집
                    url_kb3 = 'https://api.kbland.kr/land-price/price/BasePrcInfoNew?단지기본일련번호=' + str(sno) + '&면적일련번호=' + str(area_sno)
                    # print(url_kb3)

                    dict_data = {'key' : 'value'}
                    response_kb3 = requests.get(url=url_kb3, params=dict_data)
                    json_data3 = json.loads(response_kb3.text)

                    area_pyung = json_data3['dataBody']['data']['시세'][0]['공급면적평수']
                    area_supply = json_data3['dataBody']['data']['시세'][0]['공급면적']
                    area_own = json_data3['dataBody']['data']['시세'][0]['전용면적']
                    price_trade = json_data3['dataBody']['data']['시세'][0]['매매일반거래가']
                    price_high = json_data3['dataBody']['data']['시세'][0]['매매상한가']
                    price_low = json_data3['dataBody']['data']['시세'][0]['매매하한가']
                    price_avg = json_data3['dataBody']['data']['시세'][0]['매매평균가']
                    std_date = json_data3['dataBody']['data']['시세'][0]['기준년월일']

                    price_info.append([area_pyung, area_supply, area_own, price_trade, price_high, price_low, price_avg, std_date])

                self.kb_price = pd.DataFrame(price_info)

                print(self.kb_price)
                print('KB 부동산 시세정보 추출 OK')

            except:
                easygui.msgbox("시스템 에러가 발생했습니다.(cbf5) 다시 시작해 주세요", title="Error Message")

            if self.btn1_state == 1:
                self.btn2.setEnabled(True)
                self.btn2_guide.setText("클릭 ☞")                
            else :                
                self.btn2.setEnabled(False)

        else:
            pass

    def comboBoxFunction6(self):

        idx6 = self.cb6.currentIndex()
        rt_apt = self.cb6.currentText()
        print('기준년월 : ', self.year, self.quarter)
        print('RT 부동산 선택 OK : ', rt_apt)

        apt_code = self.list_rt[rt_apt][0] # 아파트 고유번호
        # print(apt_code)

        if idx6 > 0:

            try:
                self.lb6.setText('☞ 국토교통부 실거래가 아파트"를 선택해 주세요 ')

                # 아파트 실거래가 조회
                header = {
                    'Referer': 'https://rt.molit.go.kr/srh/srh.do?menuGubun=A&srhType=LOC&houseType=1&gubunCode=LAND',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36'
                        }

                rt_url2 = 'https://rt.molit.go.kr/srh/getListAjax.do'

                rt_post_data2 = {
                    'reqPage': 'SRH',
                    'menuGubun': 'A',
                    'srhType': 'LOC',
                    'houseType': 1,
                    'srhYear': self.year,
                    'srhPeriod': self.quarter,
                    'gubunCode': 'LAND',
                    'sidoCode': si_code,
                    'gugunCode': gu_code,
                    'dongCode': dong_code,
                    'danjiCode': apt_code,
                    'rentAmtType': '3'
                }

                rt_res2 = requests.get(rt_url2, rt_post_data2, headers=header)
                rt_json_data2 = json.loads(rt_res2.text)

            except:
                easygui.msgbox("시스템 에러가 발생했습니다.(cbf6:JSON) 다시 시작해 주세요", title="Error Message")

            print("RT JSON OK")

            # 크롤링 경고 메세지
            dic_key = list(rt_json_data2['jsonList'][0])
            if dic_key[0] == 'CNT':
                if rt_json_data2['jsonList'][0]['CNT'] >= 50:
                    print("크롤링 경고 : 50회 이상")
                    easygui.msgbox("국토교통부 실거래가 조회 사이트에서 보안번호를 입력 후 재시작해 주십시오.", title="크롤링 경고")
            else :
                pass
            print("크롤링 경고 OK")

            try:
                rt_apt_info = []
               
                type_n = len(rt_json_data2['jsonList'])

                n = 0
                
                for n in range(type_n)    :

                    fst_mon = len(rt_json_data2['jsonList'][n]['month1List'])
                    scd_mon = len(rt_json_data2['jsonList'][n]['month2List'])
                    thrd_mon = len(rt_json_data2['jsonList'][n]['month3List'])
    
                    cnt_trade = fst_mon + scd_mon + thrd_mon
    
                    print("lenth of jsonList :", fst_mon, scd_mon, thrd_mon, cnt_trade)

                    if len(rt_json_data2['jsonList'][n]['month1List']) > 0:
                        i9 = 0
                        for i9 in range(len(rt_json_data2['jsonList'][n]['month1List'])):
                            rt_apt_area = rt_json_data2['jsonList'][n]['month1List'][i9]['BLDG_AREA']
                            rt_apt_floor = rt_json_data2['jsonList'][n]['month1List'][i9]['APTFNO']
                            rt_apt_price = rt_json_data2['jsonList'][n]['month1List'][i9]['SUM_AMT']
                            rt_apt_mm = rt_json_data2['jsonList'][n]['month1List'][i9]['DEAL_MM']
                            rt_apt_dd = rt_json_data2['jsonList'][n]['month1List'][i9]['DEAL_DD']
    
                            rt_apt_info.append([rt_apt_area, rt_apt_mm, rt_apt_dd, rt_apt_floor, rt_apt_price])
    
                    if len(rt_json_data2['jsonList'][n]['month2List']) > 0:
                        i9_2 = 0
                        for i9_2 in range(len(rt_json_data2['jsonList'][n]['month2List'])):
                            rt_apt_area = rt_json_data2['jsonList'][n]['month2List'][i9_2]['BLDG_AREA']
                            rt_apt_floor = rt_json_data2['jsonList'][n]['month2List'][i9_2]['APTFNO']
                            rt_apt_price = rt_json_data2['jsonList'][n]['month2List'][i9_2]['SUM_AMT']
                            rt_apt_mm = rt_json_data2['jsonList'][n]['month2List'][i9_2]['DEAL_MM']
                            rt_apt_dd = rt_json_data2['jsonList'][n]['month2List'][i9_2]['DEAL_DD']
    
                            rt_apt_info.append([rt_apt_area, rt_apt_mm, rt_apt_dd, rt_apt_floor, rt_apt_price])
    
                    if len(rt_json_data2['jsonList'][n]['month3List']) > 0:
                        i9_3 = 0
                        for i9_3 in range(len(rt_json_data2['jsonList'][n]['month3List'])):
                            rt_apt_area = rt_json_data2['jsonList'][n]['month3List'][i9_3]['BLDG_AREA']
                            rt_apt_floor = rt_json_data2['jsonList'][n]['month3List'][i9_3]['APTFNO']
                            rt_apt_price = rt_json_data2['jsonList'][n]['month3List'][i9_3]['SUM_AMT']
                            rt_apt_mm = rt_json_data2['jsonList'][n]['month3List'][i9_3]['DEAL_MM']
                            rt_apt_dd = rt_json_data2['jsonList'][n]['month3List'][i9_3]['DEAL_DD']
    
                            rt_apt_info.append([rt_apt_area, rt_apt_mm, rt_apt_dd, rt_apt_floor, rt_apt_price])

                rt_price = pd.DataFrame(rt_apt_info)
                rt_price.columns=['평형','거래월','거래일','층','거래금액']
                self.rt_price = rt_price.sort_values(by=['평형','거래월','거래일','층'], axis=0)
                self.rt_apt_name = rt_json_data2['jsonList'][0]['BLDG_NM']

                print(self.rt_apt_name)
                print(self.rt_price)

            except :
                print(rt_json_data2['jsonList'][0])
                easygui.msgbox("시스템 에러가 발생했습니다.(cbf6:list) 다시 시작해 주세요", title="Error Message")

            print('RT 부동산 시세정보 추출 OK')

            if self.btn2_state == 1:
                self.btn3.setEnabled(True)
                self.btn3_guide.setText("클릭 ☞")                                
            else:
                self.btn3.setEnabled(False)

        else:
            pass

    def comboBoxFunction7(self):
        idx7 = self.cb7.currentIndex()
        self.year = self.cb7.currentText()

    def comboBoxFunction8(self):
        idx8 = self.cb8.currentIndex()
        self.quarter = self.cb8.currentText()

    def checkBoxFunction(self):

        chk_state = self.cb8_chk.isChecked()
        if self.btn2_state == 1:
            if chk_state is True :
                self.cb6.setEnabled(True)
                self.cb7.setEnabled(False)
                self.cb8.setEnabled(False)
                self.rt(si_code, gu_code, dong_code)
            else :
                self.cb6.clear()
                self.cb6.setEnabled(False)
                self.cb7.setEnabled(True)
                self.cb8.setEnabled(True)
                
        else:
            self.cb8_chk.isChecked(False)
            easygui.msgbox("KB부동산을 처리 후 입력해 주세요", title="Error Message")

    def btn1_clicked(self):
        print("부동산원 Click OK")

        # 기준일자 구하기
        std_date_source = driver2.find_element_by_xpath('//*[@id="lbAptpDt"]')        
        std_date_value = std_date_source.text.strip()
        
        # 아파트 시세정보 구하기
        area = driver2.find_elements_by_css_selector('#areaList td')

        try:
            info_list=[]
            i5 =0
            val5 = None
            info = None
            for i5, val5 in enumerate(area):
                info=str(val5.text)
                info_list.append(info)        
            apt_df = pd.DataFrame(np.array(info_list).reshape(-1,8),columns=["전용면적","세대수","매매가(만원)하한","매매가(만원)상한","전세가(만원)하한","전세가(만원) 상한","월세가(만원)보증금","월세가(만원)월세"])
            print('아파트 시세정보 추출 OK')                                                        
        except:            
            easygui.msgbox("시스템 에러가 발생했습니다.(btn1) 다시 시작해 주세요", title="Error Message")
        
        # 엑셀 저장을 위한 정보
        apt_name_find = driver2.find_element_by_xpath('//*[@id="aptName"]')
        apt_name_text = apt_name_find.text
        dt = datetime.datetime.now()
        savetime = dt.strftime("_%y_%m_%d")
        save_name = 'APT_PRICE' + '_' + str(apt_name_text) + savetime + '.xlsx'
        
        # 아파트 시세정보 엑셀 저장
        apt_df.to_excel(str(save_name), index=False)
        
        print('아파트 시세정보 엑셀 저장')                                                        
        
        # 저장된 엑셀 파일 및 시트 정보
        self.excel_filename = './' + str(save_name)
        wb = openpyxl.load_workbook(filename = self.excel_filename)    
        ws = wb.worksheets[0]
        ws.title = '한국부동산원'

        # 기준일자 엑셀 출력
        ws.insert_rows(1)
        ws.cell(1,1, std_date_value)

        # 이미지 크기 셋팅
        unit_cm1 = 12
        unit_cm2 = 12
        unit_w = round((unit_cm1/2.54)*100)    
        unit_h = round((unit_cm2/2.54)*100)    

        # 스크린캡쳐 이미지 저장
        driver2.get_screenshot_as_file('apt_capture.png')
        img1 = openpyxl.drawing.image.Image('apt_capture.png')
        img1.width = unit_w
        img1.height = unit_h
        img1.anchor = 'A20'
        ws.add_image(img1)

        # 지도 이미지 추가        
        pos3 = driver2.find_element_by_xpath('//*[@id="test_div"]/div[2]/span')
        driver2.execute_script("arguments[0].scrollIntoView(true);", pos3);

        driver2.get_screenshot_as_file('map_capture.png')
        img3 = openpyxl.drawing.image.Image('map_capture.png')
        img3.width = unit_w
        img3.height = unit_h
        img3.anchor = 'E20'
        ws.add_image(img3)

        # 실거래가 이미지 추가
        driver2.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        driver2.find_element_by_xpath('//*[@id="rtmsBtn"]').click()
        time.sleep(2)
        
        pos2 = driver2.find_element_by_xpath('//*[@id="test_div"]/div[7]')
        driver2.execute_script("arguments[0].scrollIntoView(true);", pos2);

        driver2.get_screenshot_as_file('reat_tr_capture.png')
        img2 = openpyxl.drawing.image.Image('reat_tr_capture.png')
        img2.width = unit_w
        img2.height = unit_h        
        img2.anchor = 'A42'
        ws.add_image(img2)

        # 출력 페이지 편집
        THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
        n_row = ws.max_row

        for col in [1,2,3,4,5,6,7,8]:
            for row in range(1, n_row+1):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=col).border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.column_dimensions[get_column_letter(1)].width = 10
        ws.column_dimensions[get_column_letter(2)].width = 10
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left')
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.insert_rows(1)        
        ws.insert_rows(2)                
        title = '한국부동산원 부동산 가격 조회'
        ws.cell(1, 1, title)

        # 브라우저 창크기 조절
        # driver2.set_window_size(200,200)

        # 엑셀저장
        wb.save(filename = self.excel_filename)
        wb.close()
        
        print(text1, text2, text3)
        
        self.body4.setText("■■■■□□")
        self.msg1.setText('■ 정상 처리되었습니다. 다음 단계가 활성화 되었습니다 ')

        self.btn1_state = 1

        # 비활성화
        self.cb4.setEnabled(False)
        self.btn1.setEnabled(False)

        # 활성화
        self.cb5.setEnabled(True)        

        print("한국부동산원 처리 OK")

        # 에러처리용A : 법정동 코드 전달
        self.kb(dong_code)

        # 에러처리용B
        # self.kb()

    def btn2_clicked(self):
        print('KB선택 OK')
        wb = openpyxl.load_workbook(filename = self.excel_filename)    
        if 'KB' in wb.sheetnames:
            pass
        else:    
            wb.create_sheet('KB', 1)
        
        ws2 = wb['KB']
       
        # 아파트 시세정보 엑셀 저장
        rows= dataframe_to_rows(self.kb_price)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
                
        print('KB부동산 시세정보 엑셀 저장')
                                                  
        # 스크린캡쳐 이미지 추가        
        unit_cm1 = 16
        unit_cm2 = 12
        unit_w = round((unit_cm1/2.54)*100)    
        unit_h = round((unit_cm2/2.54)*100)    
        
        driver3.get_screenshot_as_file('KB_apt_capture.png')
        img4 = openpyxl.drawing.image.Image('KB_apt_capture.png')
        img4.width = unit_w
        img4.height = unit_h
        img4.anchor = 'A24'
        ws2.add_image(img4)        

        # 출력 페이지 편집
        ws2.delete_rows(1)
        ws2.delete_cols(1)

        THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
        n_row = ws2.max_row

        for col in [1,2,3,4,5,6,7,8]:
            for row in range(1, n_row+1):
                ws2.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                ws2.cell(row=row, column=col).border = THIN_BORDER
            ws2.column_dimensions[get_column_letter(col)].width = 12
        ws2.column_dimensions[get_column_letter(1)].width = 15
        ws2.column_dimensions[get_column_letter(4)].width = 15
        ws2.sheet_properties.pageSetUpPr.fitToPage = True

        ws2.insert_rows(1)
        title = 'KB 부동산 가격 조회'
        ws2.cell(1, 1, title)

        # 브라우저 창크기 조절
        # driver3.set_window_size(200,200)

        wb.save(filename = self.excel_filename)
        wb.close()
        self.body4.setText("■■■■■□")
        self.msg2.setText('■ 정상 처리되었습니다. 마지막 단계가 활성화 되었습니다 ')

        self.btn2_state = 1

        # 비활성화
        self.cb5.setEnabled(False)
        self.btn2.setEnabled(False)

        # 활성화
        # self.cb6.setEnabled(True)
        self.cb7.setEnabled(True)
        self.cb8.setEnabled(True)

        print("KB 부동산 처리 OK")

        # 실거래가용 변수 전달
        self.year = datetime.datetime.now().year
        self.quarter = (datetime.datetime.now().month - 1) // 3 + 1
        # self.rt(si_code, gu_code, dong_code)

    def btn3_clicked(self):

        wb = openpyxl.load_workbook(filename=self.excel_filename)
        if '실거래가' in wb.sheetnames:
            pass
        else:
            wb.create_sheet('실거래가', 2)

        ws3 = wb['실거래가']

        # 아파트 시세정보 엑셀 저장
        rows = dataframe_to_rows(self.rt_price)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=value)
        print('RT 부동산 시세정보 엑셀 저장')

        # 출력 페이지 편집

        ws3.delete_cols(1)

        THIN_BORDER = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
        n_row = ws3.max_row

        for col in [1, 2, 3, 4, 5]:
            for row in range(1, n_row + 1):
                ws3.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                ws3.cell(row=row, column=col).border = THIN_BORDER
            ws3.column_dimensions[get_column_letter(col)].width = 12
        ws3.sheet_properties.pageSetUpPr.fitToPage = True

        # 타이틀 및 기본정보, 기준일자 출력
        ws3.insert_rows(1)
        ws3.insert_rows(2)
        ws3.insert_rows(3)
        ws3.insert_rows(4)
        ws3.delete_rows(6)        
        title = '국토교통부 실거래가 조회'

        std_date_value2 = str(self.year) + "년  " + str(self.quarter) + "분기"

        ws3.cell(1, 1, title)
        ws3.cell(2, 1, std_date_value2)
        ws3.cell(3, 1, self.rt_apt_name)

        # 브라우저 창크기 조절
        # driver3.set_window_size(200,200)

        wb.save(filename=self.excel_filename)
        wb.close()

        self.body4.setText("■■■■■■")
        self.msg3.setText('■ 정상 처리되었습니다. 엑셀파일을 확인해 주세요')

        # 비활성화
        self.cb6.setEnabled(False)
        self.cb7.setEnabled(False)
        self.cb8.setEnabled(False)
        self.btn3.setEnabled(False)
        self.cb8_chk.setEnabled(False)

        self.body5.setText('    ☞ 미션 성공!! 조회가 완료되었습니다. ^^')

        print("OK")

    def btn4_clicked(self):

        self.cb1.clear()
        self.cb2.clear()
        self.cb3.clear()
        self.cb4.clear()
        self.cb5.clear()
        self.cb6.clear()
        self.cb7.clear()
        self.cb8.clear()

        self.btn1_state = 0
        self.btn2_state = 0

        self.cb4.setEnabled(False)
        self.cb5.setEnabled(False)
        self.cb6.setEnabled(False)

        self.lb1.setText('가. 시 도 선택')
        self.lb2.setText('나. 구 군 선택')
        self.lb3.setText('다. 읍면동 선택')
        self.lb4.setText('라. 아파트 선택 : 한국부동산원')
        self.lb5.setText('마. 아파트 선택 : KB 국민은행')
        self.lb6.setText('마. 아파트 선택 : 국토교통부 실거래가')

        self.body4.setText('□□□□□□')

        self.msg1.setText('＃ 한국부동산원 처리 후 아래 KB 부동산 선택이 가능합니다')
        self.msg2.setText('＃ KB부동산 처리 후 아래 실거래가 선택이 가능합니다')
        self.msg3.setText('＃ 아직 처리가 완료되지 않았습니다')

        self.close()

        self.__init__()
        
    def btn5_clicked(self):
        
        self.close()

if __name__ == '__main__':

    if datetime.datetime.today() > datetime.datetime(2022,12,31) :
        tk.messagebox.showinfo("경고","프로그램 유효기간이 지났습니다. 개발자에게 문의해 주세요")
        sys.exit()

    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    driver2 = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    driver3 = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    driver4 = webdriver.Chrome(ChromeDriverManager().install(), options=options)

    adr = 'http://www.rtech.or.kr/rtech/main/mapSearch.do?#'
    driver.get(adr)
    
    adr3 = 'https://kbland.kr/map?xy=37.5205559,126.9265729,16'
    driver3.get(adr3)

    rt_adr = 'https://rt.molit.go.kr/srh/srh.do?menuGubun=A&srhType=LOC&houseType=1&gubunCode=LAND'
    driver4.get(rt_adr)

    app = QApplication(sys.argv)

    a = UI_FORM()

    sys.exit(app.exec_())


# *********************************************

